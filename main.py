import argparse
import os.path
from PyPDF2 import PdfFileReader
import re
from openpyxl import Workbook
from openpyxl.styles import Font

title = Font(name='Calibri', size=11, bold=True)
header = Font(name='Calibri', size=12, bold=True, underline='single')

graph = {
    0: "Absatz Markt 1",
    1: "Preis Markt 1",
    2: "Marktanteil",
    3: "Preisrange Markt 1",
    4: "Werbebudget",
    5: "in %",
    6: "Bewertung Werbung",
    7: "Vertriebsmitarbeiter",
    8: "Qualifikationsniveau",
    9: "Vertriebsmit Branche",
    10: "Vertriebsmit Quali Branche",
    11: "Bekanntheitsindex",
    12: "Prozent Kunden",
    13: "Bewertung Bekanntheit",
    14: "Kundenzufriedenheit",
    15: "Kundenzufriedenheit Branche",
    16: "Prozessoptimierung",
    17: "Bewertung Prozessoptimierung",
    18: "Produktivitätsindex",
    19: "Innovationsindex",
    20: "Innovationsindex Branche",
    21: "pot. Absatz Markt 1",
    22: "Umsatz Markt 1",
    23: "Verwaltungsmitarbeiter",
    24: "Entwickler:innen"

}

graph_period_four = {
    0: "Absatz Markt 1",
    1: "Preis Markt 1",
    2: "Marktanteil",
    3: "Preisrange Markt 1",
    4: "Absatz Markt 2",
    5: "Preisrange Markt 2",
    6: "Werbebudget",
    7: "in %",
    8: "Bewertung Werbung",
    9: "Vertriebsmitarbeiter",
    10: "Qualifikationsniveau",
    11: "Vertriebsmit Branche",
    12: "Vertriebsmit Quali Branche",
    13: "Bekanntheitsindex",
    14: "Prozent Kunden",
    15: "Bewertung Bekanntheit",
    16: "Kundenzufriedenheit",
    17: "Kundenzufriedenheit Branche",
    18: "Prozessoptimierung",
    19: "Bewertung Prozessoptimierung",
    20: "Produktivitätsindex",
    21: "Innovationsindex",
    22: "Innovationsindex Branche",
    23: "pot. Absatz Markt 1",
    24: "Umsatz Markt 1",
    25: "pot. Absatz Markt 2",
    26: "Preis Markt 2",
    27: "Umsatz Markt 2",
    28: "Verwaltungsmitarbeiter",
    29: "Entwickler:innen"

}

d = {
    "Umsatzrendite": "",
    "Absatz Markt 1": [],
    "pot. Absatz Markt 1": [],
    "Preis Markt 1": [],
    "Marktanteil": [],
    "Preisrange Markt 1": [],
    "Umsatz Markt 1": [],
    "Absatz Markt 2": [],
    "pot. Absatz Markt 2": [],
    "Preis Markt 2": [],
    "Preisrange Markt 2": [],
    "Umsatz Markt 2": [],
    "Verwaltungsmitarbeiter": [],
    "Kommunikationspolitik": "",
    "Werbebudget": [],
    "in %": [],
    "Bewertung Werbung": [],
    "Bekanntheitsindex": [],
    "Prozent Kunden": [],
    "Bewertung Bekanntheit": [],
    "Distributionspolitik": "",
    "Vertriebsmitarbeiter": [],
    "Vertriebsmit Branche": [],
    "Qualifikationsniveau": [],
    "Vertriebsmit Quali Branche": [],
    "Kundenzufriedenheit": [],
    "Kundenzufriedenheit Branche": [],
    "Prozessoptimierung, Training": "",
    "Prozessoptimierung": [],
    "Bewertung Prozessoptimierung": [],
    "Entwickler:innen": [],
    "Produktivitätsindex": [],
    "Innovationsindex": [],
    "Innovationsindex Branche": [],
}
regex_site_one = r".*?(\d+\..{3}).*?(\d{3}).*?(\d{2} %).*?(\d{3} und \d{3}).*?(\d{2,3}\.\d{3}).*?(\d{2} \%).*?(\w*) " \
                 r"W.*?(\d{1,2}).*?(\d{2,3}).*?(\d{1,2} und \d{1,2}).*?(\d{2,3} und \d{2,3}).*?(\d{2," \
                 r"3}).*?(\d{2}).*?(\w*) Bekann.*?(\d{2,3}).*?(\d{2,3} bis \d{2,3}).*?(\d{2,3}.\d{3}).*tz (\w* " \
                 r"?\w*).*?(\d{2,3}).*?(\d{2,3}).*?(\d{2,3}.*?\d{2,3}) "
regex_site_one_markt_two = r".*?(\d+\..{3}).*?(\d{3}).*?(\d{2} %).*?(\d{3} und \d{3}).*?(\d{0,2}\.?\d{3}).*?(\d{3} " \
                           r"und \d{3}).*?(\d{2,3}\.\d{3}).*?(\d{2} \%).*?(\w*) W.*?(\d{1,2}).*?(\d{2,3}).*?(\d{1," \
                           r"2} und \d{1,2}).*?(\d{2,3} und \d{2,3}).*?(\d{2,3}).*?(\d{2}).*?(\w*) Bekann.*?(\d{2," \
                           r"3}).*?(\d{2,3} bis \d{2,3}).*?(\d{2,3}.\d{3}).*tz (\w* ?\w*).*?(\d{2,3}).*?(\d{2," \
                           r"3}).*?(\d{2,3}.*?\d{2,3}) "
regex_site_three_p1 = r".*?po.*?t.*?\d.*?(\d+\.\d{3}).*?Um.*?\d.*?(\d?\.?\d{1,3}\.\d{3}).*?End.*?(\d) (\d)"
regex_site_three = r".*?po.*?t.*?\d.*?\d{1,}.\d{3} (\d{1,}.\d{3}).*?Um.*?\d.*?\d?.?\d{1,}.\d{3}.\d{3} (\d*.?\d{2," \
                   r"}.\d{3}).*?End.*?(\d) (\d) "
regex_site_three_markt_two = r".*?po.*?t 1.*?\d{1,2}.\d{3} (\d{1,2}\.\d{3}).*?Umsatz.*?\d.*?\d.\d{3}.\d{3} (\d.\d{" \
                             r"3}.\d{3}).*?po.*?\d?.?\d{3} (\d?.?\d{3}).*?Preis.*?\d.*?\d{3} (\d{3}).*?Um.*?\d.*?\d{" \
                             r"1,3}.\d{3} (\d{1,3}.\d{3}).*?End.*?(\d) (\d) "


def fill_empty():
    d["Umsatz Markt 2"].append(0)
    d["pot. Absatz Markt 2"].append(0)
    d["Absatz Markt 2"].append(0)
    d["Preis Markt 2"].append(0)
    d["Preisrange Markt 2"].append(0)


if __name__ == "__main__":

    name = ""
    directory = ""
    start = 1
    end = 8

    parser = argparse.ArgumentParser()
    parser.add_argument('-n', help="Name of the Project")
    parser.add_argument('-d', help="Data Directory")
    parser.add_argument('-s', help="Start Period")
    parser.add_argument('-e', help="End Period")
    args = parser.parse_args()
    print(args)

    if args.d:
        directory = args.d
    if args.s:
        start = args.s
    if args.End:
        end = args.e
    if args.Name:
        name = args.n
    else:
        name = os.path.basename(directory)

    print(start)
    print(end)
    print(directory)
    print(name)

    for i in range(start, end + 1):
        with open(f'{directory}/reportP{i}U0.pdf', 'rb') as f:
            pdf = PdfFileReader(f)
            first_page = pdf.getPage(0)
            first_page = re.sub(r'\s{2,}', ' ', first_page.extractText().replace('\n', ' '))

            third_page = pdf.getPage(2)
            third_page = re.sub(r'\s{2,}', ' ', third_page.extractText().replace('\n', ' '))

            if i <= 3 or i == 8:
                groups_one = re.match(regex_site_one, first_page).groups()
                groups_three = re.match(regex_site_three_p1, third_page).groups()
                currentgraph = graph
                fill_empty()

            elif i > 3:
                groups_one = re.match(regex_site_one_markt_two, first_page).groups()
                groups_three = re.match(regex_site_three_markt_two, third_page).groups()
                currentgraph = graph_period_four

            groups = groups_one + groups_three

            for group, index in zip(groups, range(len(currentgraph))):
                group = group.replace('.', '')
                if group.isdigit():
                    d[currentgraph[index]].append(int(group))
                else:
                    d[currentgraph[index]].append(group)

    wb = Workbook()
    ws = wb.active
    ws['A1'] = name
    ws['A1'].font = header

    row_start = 2
    colum_start = 1

    for key, value in d.items():
        ws.cell(row=row_start, column=colum_start, value=key)
        if value == "":
            # Style cell as title
            ws.cell(row=row_start, column=colum_start).font = title

        for i in range(len(value)):
            ws.cell(row=row_start, column=colum_start + i + 1, value=value[i])
        row_start += 1

    try:
        wb.save(f'{directory}/{name}.xlsx')
    except PermissionError:
        print("Please close the file")
        input("Press Enter to continue...")
        wb.save(f'{directory}/{name}.xlsx')
